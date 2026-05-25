"""表格数据服务"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import BusinessError
from app.modules.tabular_data.repository import TabularJobRepository


class TabularDataService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = TabularJobRepository(db)

    def _allowed_extensions(self) -> set[str]:
        raw = settings.tabular_allowed_extensions
        return {item.strip().lower() for item in raw.split(",") if item.strip()}

    def _validate_upload(self, filename: str, content: bytes):
        extension = Path(filename or "").suffix.lower().lstrip(".")
        if extension not in self._allowed_extensions():
            raise BusinessError("file extension is not allowed", status_code=400)

        max_size = settings.max_upload_size_mb * 1024 * 1024
        if len(content) > max_size:
            raise BusinessError("file exceeds max upload size", status_code=400)

        if not content:
            raise BusinessError("file is empty", status_code=400)

    def _parse_csv_preview(self, content: bytes) -> dict:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        headers = reader.fieldnames or []
        rows = [dict(row) for row in reader]
        return {
            "headers": headers,
            "rows": rows,
            "sample_rows": rows[:5],
            "row_count": len(rows),
        }

    def _parse_xlsx_preview(self, content: bytes) -> dict:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise BusinessError("xlsx import requires openpyxl", status_code=500) from exc

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        first_row = next(rows_iter, None)
        if first_row is None:
            return {"headers": [], "rows": [], "sample_rows": [], "row_count": 0}

        headers = [str(item or "") for item in first_row]
        rows = []
        for value_row in rows_iter:
            row_dict = {
                headers[i]: value_row[i] if i < len(value_row) else "" for i in range(len(headers))
            }
            rows.append(row_dict)

        return {
            "headers": headers,
            "rows": rows,
            "sample_rows": rows[:5],
            "row_count": len(rows),
        }

    def preview_import(self, filename: str, content: bytes) -> dict:
        self._validate_upload(filename, content)

        extension = Path(filename).suffix.lower()
        if extension == ".csv":
            parsed = self._parse_csv_preview(content)
        elif extension == ".xlsx":
            parsed = self._parse_xlsx_preview(content)
        else:
            raise BusinessError("file extension is not allowed", status_code=400)

        return {
            "headers": parsed["headers"],
            "sample_rows": parsed["sample_rows"],
            "row_count": parsed["row_count"],
        }

    def create_export_job(self, payload: dict) -> dict:
        job = self.repo.create_export_job(created_by="system", params_json=payload)
        self.db.commit()

        # 简化: 实际应异步执行
        job.status = "SUCCESS"
        self.db.commit()

        return {"job_id": job.id, "status": job.status}

    def get_export_job(self, job_id: int) -> dict:
        job = self.repo.get_by_id(job_id)
        if job is None:
            raise BusinessError("job not found", status_code=404)

        return {
            "job_id": job.id,
            "status": job.status,
            "error_message": job.error_message,
        }
